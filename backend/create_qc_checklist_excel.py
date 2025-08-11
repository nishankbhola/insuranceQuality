"""
Create Excel file with Application QC Checklist for review and modification
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_qc_checklist_excel():
    """Create Excel file with current QC checklist"""
    
    # Current QC checklist data
    checklist_data = [
        # Forms
        {
            "ID": "privacy_consent",
            "Category": "Forms",
            "Checklist Item": "Privacy Consent Form signed",
            "Current Logic": "Currently placeholder - needs implementation",
            "Application Data Used": "Placeholder",
            "Quote Data Used": "Not used",
            "Current Status": "Critical (Required=True)",
            "Keep/Remove/Modify": "REVIEW",
            "Your Comments": "",
            "New Logic (if different)": ""
        },
        
        # Other Requirements
        {
            "ID": "payment_method_valid",
            "Category": "Other Requirements",
            "Checklist Item": "Valid payment method provided",
            "Current Logic": "Check if payment_frequency field exists",
            "Application Data Used": "policy_info.payment_frequency",
            "Quote Data Used": "Not used",
            "Current Status": "Critical (Required=True)",
            "Keep/Remove/Modify": "REVIEW",
            "Your Comments": "",
            "New Logic (if different)": ""
        }
    ]
    
    # Create DataFrame
    df = pd.DataFrame(checklist_data)
    
    # Create Excel workbook with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "QC Checklist Review"
    
    # Add headers
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
    
    # Add data
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )
            
            # Color code based on category
            if col_num == 2:  # Category column
                if value == "Completed Information":
                    cell.fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
                elif value == "Driver/MVR":
                    cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                elif value == "Coverage Requirements":
                    cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
                elif value == "Forms":
                    cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                elif value == "Other Requirements":
                    cell.fill = PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid")
            
            # Highlight Keep/Remove/Modify column
            if col_num == 8:  # Keep/Remove/Modify column
                cell.fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
                cell.font = Font(bold=True, color="FF6600")
            
            # Highlight Your Comments column
            if col_num == 9:  # Your Comments column
                cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                cell.font = Font(color="0066CC")
    
    # Auto-adjust column widths
    column_widths = {
        'A': 20,  # ID
        'B': 18,  # Category
        'C': 35,  # Checklist Item
        'D': 40,  # Current Logic
        'E': 30,  # Application Data Used
        'F': 25,  # Quote Data Used
        'G': 20,  # Current Status
        'H': 15,  # Keep/Remove/Modify
        'I': 40,  # Your Comments
        'J': 40   # New Logic
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Set row heights
    for row in range(2, len(checklist_data) + 2):
        ws.row_dimensions[row].height = 60
    
    # Add instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ["INSTRUCTIONS FOR REVIEWING QC CHECKLIST", ""],
        ["", ""],
        ["1. Review Each Item:", "Look at each QC checklist item and its current logic"],
        ["", ""],
        ["2. In 'Keep/Remove/Modify' column, specify:", ""],
        ["   - KEEP", "Keep the item as-is"],
        ["   - REMOVE", "Remove this item from checklist"],
        ["   - MODIFY", "Keep item but change the logic"],
        ["", ""],
        ["3. In 'Your Comments' column:", "Add your feedback, business rules, or concerns"],
        ["", ""],
        ["4. In 'New Logic' column:", "If MODIFY, describe the new logic you want"],
        ["", ""],
        ["Examples:", ""],
        ["", ""],
        ["Keep/Remove/Modify: MODIFY", ""],
        ["Your Comments: Date should allow 30-day variance, not exact match", ""],
        ["New Logic: Allow application date within 30 days of effective date", ""],
        ["", ""],
        ["Keep/Remove/Modify: REMOVE", ""],
        ["Your Comments: We don't use OPCF 43 in our business", ""],
        ["", ""],
        ["Keep/Remove/Modify: KEEP", ""],
        ["Your Comments: This is correct as-is", ""],
        ["", ""],
        ["CATEGORIES EXPLAINED:", ""],
        ["", ""],
        ["Completed Information:", "Ensuring all required fields are filled"],
        ["Driver/MVR:", "Validating driver information and MVR consistency"],
        ["Coverage Requirements:", "Checking coverage limits and endorsements"],
        ["Forms:", "Validating required forms are complete"],
        ["Other Requirements:", "Payment methods, broker info, etc."]
    ]
    
    for row_num, (col1, col2) in enumerate(instructions, 1):
        ws2.cell(row=row_num, column=1, value=col1).font = Font(bold=True if col1 and not col1.startswith(" ") else False)
        ws2.cell(row=row_num, column=2, value=col2)
    
    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['B'].width = 50
    
    # Save the file
    filename = "QC_Checklist_Review.xlsx"
    wb.save(filename)
    
    print(f"✅ Excel file created: {filename}")
    print(f"📊 Contains {len(checklist_data)} QC checklist items")
    print(f"📝 Please review and update the 'Keep/Remove/Modify' and 'Your Comments' columns")
    
    # Also create a CSV version
    csv_filename = "QC_Checklist_Review.csv"
    df.to_csv(csv_filename, index=False)
    print(f"📋 CSV file also created: {csv_filename}")
    
    return filename, csv_filename

if __name__ == "__main__":
    create_qc_checklist_excel()

